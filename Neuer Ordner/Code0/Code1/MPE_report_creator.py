################################################################################
########################## MPE report creator #################################
################################################################################
################################################################################
################################################################################
################################################################### Min ########
################################################################################

# basic package for data analysis
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import time

#### for Boxcox limit ####
#from scipy.stats import anderson
#from scipy import stats

# read excel file (already cleaned from SQL)
df = pd.read_excel('D:\\books\\MPE_project\\Statistics\\NY_bin_statistics.xlsx')

# data cleaning for analysis - new column for lot id and remove empty and contact failure
df['lot_ID'] = df['lot'].map(str) + '_' + df['sub_lot'].map(str)
df = df[df['bin_name'] != 'Leer']
df = df[df['bin_name'] != 'Kelvin']

# percentage calculation
df_indivisual_count = df.groupby(['lot_ID','bin_number', 'bin_name']).agg({'count': sum})
df_percentage = df_indivisual_count.groupby(level = 0).apply(lambda x : 100 * x / float(x.sum()))

# index reset for order
df_percentage = df_percentage.reset_index()

# change name
df_percentage = df_percentage.rename(columns = {'count':'percentage'})

# extract unique lot ID
unique_lot_ID = list(set(df_percentage['lot_ID']))
unique_lot_ID = sorted(unique_lot_ID)

# devide dataframe into three data
df_percentage_1 = df_percentage[:33] # # of tests 11
df_percentage_2 = df_percentage[33:585] # # of tests 12
df_percentage_3 = df_percentage[585:] # # of tests 14

# devide dataframe into three data for date
df_date_1 = df[:33] # # of tests 11
df_date_2 = df[33:585] # # of tests 12
df_date_3 = df[585:] # # of tests 14

# current date
ww = time.strftime("%W")
yy = time.strftime("%Y")

# open the format for MPE report
# MPE = openpyxl.load_workbook(filename='C:\\Users\\minsung_kim\\books\\MPE_project\\MPE_FMMT614-7-55_format - Copy.xlsx')
MPE = openpyxl.load_workbook(filename='D:\\books\\MPE_project\\MPE_FMMT614-7-55_format.xlsx')

# grab a sheet in the excel file
ws1 = MPE["Tabelle1"]

# first three lots without Rth
li = 0
for i in range(1, int(len(df_percentage_1)/11) + 1):
    c1 = 0 # lot id and yield cycle
    c2 = 4 # open -
    c3 = 5 # short -
    c4 = 6 # ICBO -
    c5 = 7 # ICES -
    c6 = 8 # IEBO -
    c7 = 9 # VCEsat -
    c8 = 10 # VBEsat -
    ws1.cell(row = li + 14, column = 1).value = df_percentage_1.iloc[c1 + (i-1)*11][0] # 1 lot id
    ws1.cell(row = li + 14, column = 2).value = df_date_1.iloc[c1 + (i-1)*11][5] # date
    ws1.cell(row = li + 14, column = 3).value = df_percentage_1.iloc[c1 + (i-1)*11][3] # 3 yield
    ws1.cell(row = li + 14, column = 4).value = df_percentage_1.iloc[c2 + (i-1)*11][3] # 4 open
    ws1.cell(row = li + 14, column = 5).value = df_percentage_1.iloc[c3 + (i-1)*11][3] # 5 short
    ws1.cell(row = li + 14, column = 6).value = df_percentage_1.iloc[c4 + (i-1)*11][3] # 6 ICBO
    ws1.cell(row = li + 14, column = 7).value = df_percentage_1.iloc[c5 + (i-1)*11][3] # 7 ICES
    ws1.cell(row = li + 14, column = 8).value = df_percentage_1.iloc[c6 + (i-1)*11][3] # 8 IEBO
    ws1.cell(row = li + 14, column = 9).value = df_percentage_1.iloc[c7 + (i-1)*11][3] # 9 VCEsat
    ws1.cell(row = li + 14, column = 10).value = df_percentage_1.iloc[c8 + (i-1)*11][3] # 10 VBEsat
    li = li + 1

# second 46 lots with Rth
for i in range(1, int(len(df_percentage_2)/12) + 1):
    c1 = 0 # lot id and yield cycle
    c2 = 4 # open -
    c3 = 5 # short -
    c4 = 6 # ICBO -
    c5 = 7 # ICES -
    c6 = 8 # IEBO -
    c7 = 9 # VCEsat -
    c8 = 10 # VBEsat -
    c9 = 11 # Rth -
    ws1.cell(row = li + 14, column = 1).value = df_percentage_2.iloc[c1 + (i-1)*12][0] # 1 lot id
    ws1.cell(row = li + 14, column = 2).value = df_date_2.iloc[c1 + (i-1)*12][5] # date
    ws1.cell(row = li + 14, column = 3).value = df_percentage_2.iloc[c1 + (i-1)*12][3] # 3 yield
    ws1.cell(row = li + 14, column = 4).value = df_percentage_2.iloc[c2 + (i-1)*12][3] # 4 open
    ws1.cell(row = li + 14, column = 5).value = df_percentage_2.iloc[c3 + (i-1)*12][3] # 5 short
    ws1.cell(row = li + 14, column = 6).value = df_percentage_2.iloc[c4 + (i-1)*12][3] # 6 ICBO
    ws1.cell(row = li + 14, column = 7).value = df_percentage_2.iloc[c5 + (i-1)*12][3] # 7 ICES
    ws1.cell(row = li + 14, column = 8).value = df_percentage_2.iloc[c6 + (i-1)*12][3] # 8 IEBO
    ws1.cell(row = li + 14, column = 9).value = df_percentage_2.iloc[c7 + (i-1)*12][3] # 9 VCEsat
    ws1.cell(row = li + 14, column = 10).value = df_percentage_2.iloc[c8 + (i-1)*12][3] # 10 VBEsat
    ws1.cell(row = li + 14, column = 12).value = 0.01*df_percentage_2.iloc[c9 + (i-1)*12][3] # 12 Rth
    li = li + 1

# current third lots
for i in range(1, int(len(df_percentage_3)/14) + 1):
    c1 = 0 # lot id and yield cycle 1
    c2 = 1 #  yield cycle 2
    c3 = 2 #  yield cycle 3
    c4 = 6 # open -
    c5 = 7 # short -
    c6 = 8 # ICBO -
    c7 = 9 # ICES -
    c8 = 10 # IEBO -
    c9 = 11 # VCEsat -
    c10 = 12 # VBEsat -
    c11 = 13 # Rth -
    ws1.cell(row = li + 14, column = 1).value = df_percentage_3.iloc[c1 + (i-1)*14][0] # 1 lot id
    ws1.cell(row = li + 14, column = 2).value = df_date_3.iloc[c1 + (i-1)*14][5] # date
    ws1.cell(row = li + 14, column = 3).value = df_percentage_3.iloc[c1 + (i-1)*14][3] + df_percentage_3.iloc[c2 + (i-1)*14][3] + df_percentage_3.iloc[c3 + (i-1)*14][3] # 1,2,3 yield
    ws1.cell(row = li + 14, column = 4).value = df_percentage_3.iloc[c4 + (i-1)*14][3] # 4 open
    ws1.cell(row = li + 14, column = 5).value = df_percentage_3.iloc[c5 + (i-1)*14][3] # 5 short
    ws1.cell(row = li + 14, column = 6).value = df_percentage_3.iloc[c6 + (i-1)*14][3] # 6 ICBO
    ws1.cell(row = li + 14, column = 7).value = df_percentage_3.iloc[c7 + (i-1)*14][3] # 7 ICES
    ws1.cell(row = li + 14, column = 8).value = df_percentage_3.iloc[c8 + (i-1)*14][3] # 8 IEBO
    ws1.cell(row = li + 14, column = 9).value = df_percentage_3.iloc[c9 + (i-1)*14][3] # 9 VCEsat
    ws1.cell(row = li + 14, column = 10).value = df_percentage_3.iloc[c10 + (i-1)*14][3] # 10 VBEsat
    ws1.cell(row = li + 14, column = 12).value = 0.01*df_percentage_3.iloc[c11 + (i-1)*14][3] # 12 Rth
    li = li + 1

# change the date (calendar week and year) on the sheet
MPE.worksheets[0].cell(row=2, column=2).value='WW'+ ww + ' '+ yy

#save it as an excel file for DZNG
MPE.save(r'D:\\books\\MPE_project\\Report_MPE\\372S00032_FMMT614-7-55_MPE_' + yy + '_WW' + ww + '_NAT' + '.xlsx')

# open the format for MPE report
# MPE2 = openpyxl.load_workbook(filename='C:\\Users\\minsung_kim\\books\\MPE_project\\Report_MPE\\372S00032_FMMT614-7-55_MPE_' + yy + '_WW' + ww + '_NAT' + '.xlsx')
MPE2 = openpyxl.load_workbook(filename='D:\\books\\MPE_project\\Report_MPE\\372S00032_FMMT614-7-55_MPE_' + yy + '_WW' + ww + '_NAT' + '.xlsx')

# grab a sheet in the excel file
ws2 = MPE2["Tabelle1"]

# delete the column for Rth
ws2.delete_cols(12, 10) # .delete_cols(starting column, how many cols from the starting point)


#save it as an excel file for MPE
MPE2.save(r'D:\\books\\MPE_project\\Report_MPE\\372S00032_FMMT614-7-55_MPE_' + yy + '_WW' + ww + '.xlsx')
